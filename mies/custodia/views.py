from django.shortcuts import render
from django.urls import reverse_lazy
from django.contrib.auth.mixins import PermissionRequiredMixin,LoginRequiredMixin
from django.views.generic import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import (CreateView, UpdateView, DeleteView)
from .models import Custodia
from .forms import CustodiaForm
from inventario.models import Dispositivo
#excel
#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
import csv
from csv import reader
from openpyxl.utils import get_column_letter

# Create your views here.
class CustodiaListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'custodia.view_custodia'
    model = Custodia
    template_name = "custodia/custodia_listado.html"

    
    
# excel

#Nuestra clase hereda de la vista genérica TemplateView
class ComputadorasExcelView(TemplateView):
     
     
      #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las lista de nuestra base de listas

        lista = Custodia.objects.filter(estado=1)
        
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE lista'
        ws['B1'] = 'INVENTARIO  DE COMPUTADORAS MIES'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'NOMBRE DEL CUSTODIO'
        ws['C3'] = 'UNIDAD EJECUTORA'
        ws['D3'] = 'DEPENDENCIA'
        ws['E3'] = 'UNIDAD ADMINISTRATIVA'
        ws['F3'] = 'NOMBRE DEL EQUIPO'
        ws['G3'] = 'TIPO DE EQUIPO'
        ws['H3'] = 'MARCA'
        ws['I3'] = 'MODELO'
        ws['J3'] = 'SERIE'
        ws['K3'] = 'CÓDIGO MIES'
        ws['L3'] = 'PROCESADOR'
        ws['M3'] = 'CAPACIDAD DISCO DURO'
        ws['N3'] = 'TAMAÑO MEMORIA RAM'
        ws['O3'] = 'SISTEMA OPERATIVO'
        ws['P3'] = 'DIRECCIÓN IP'
        ws['Q3'] = 'DIRECCIÓN MAC'
        ws['R3'] = 'SOFTWARE OFIMÁTICA'
        ws['S3'] = 'SOFTWARE ANTIVIRUS'
        ws['T3'] = 'SOFTWARE ADICIONAL'
        ws['U3'] = 'ESTADO DEL EQUIPO'
        ws['V3'] = 'OBSERVACIONES'
       

   #     ws.merge_cells('R3:Y3')
        #ancho de columna
        ws.column_dimensions['B'].width = 40.0
        ws.column_dimensions['C'].width = 20.0
        ws.column_dimensions['D'].width = 35.0
        ws.column_dimensions['E'].width = 35.0
        ws.column_dimensions['F'].width = 30.0
        ws.column_dimensions['G'].width = 30.0
        ws.column_dimensions['H'].width = 25.0
        ws.column_dimensions['I'].width = 25.0
        ws.column_dimensions['J'].width = 25.0
        ws.column_dimensions['K'].width = 25.0
        ws.column_dimensions['L'].width = 25.0
        ws.column_dimensions['M'].width = 25.0
        ws.column_dimensions['N'].width = 25.0
        ws.column_dimensions['O'].width = 25.0
        ws.column_dimensions['P'].width = 25.0
        ws.column_dimensions['Q'].width = 25.0
        ws.column_dimensions['R'].width = 25.0
        ws.column_dimensions['S'].width = 25.0
        ws.column_dimensions['T'].width = 40.0
        ws.column_dimensions['U'].width = 25.0
        ws.column_dimensions['V'].width = 40.0
       
        cont=4

        #Recorremos el conjunto de lista y vamos escribiendo cada uno de los listas en las celdas
        
        for dato in lista:
            
            if dato.equipo.codigoMies is None:
                VcodMies= "N/A"
            else:
                VcodMies= dato.equipo.codigoMies
            if dato.equipo.serie is None:
                VserMies= "N/A"
            else:
                VserMies= dato.equipo.serie

            if dato.equipo.direccionIp is None:
                VdireccionIp= "N/A"
            else:
                VdireccionIp= dato.equipo.direccionIp

            if dato.equipo.direccionMac is None:
                VdireccionMac= "N/A"
            else:
                VdireccionMac= dato.equipo.direccionMac
           


            if dato.equipo.tipoEquipo is None:
                VtipoEquipo= "N/A"
            else:
                VtipoEquipo= str(dato.equipo.tipoEquipo)


            if dato.equipo.marca is None:
                VMarca= "N/A"
            else:
                VMarca= str(dato.equipo.marca)


            if dato.equipo.modelo is None:
                Vmodelo= "N/A"
            else:
                Vmodelo= str(dato.equipo.modelo)


            if dato.equipo.capacidadProcesador is None:
                Vprocesador= "N/A"
            else:
                Vprocesador= str(dato.equipo.capacidadProcesador)

            

            if dato.equipo.capacidadDisco is None:
                VcapacidadDisco= "N/A"
            else:
                VcapacidadDisco= str(dato.equipo.capacidadDisco)

            if dato.equipo.capacidadMemoriaRam is None:
                VcapacidadMemoriaRam= "N/A"
            else:
                VcapacidadMemoriaRam= str(dato.equipo.capacidadMemoriaRam)

         
            if dato.equipo.sistemaOperativo is None:
                VsistemaOperativo= "N/A"
            else:
                VsistemaOperativo= str(dato.equipo.sistemaOperativo)


            if dato.equipo.ofimatica is None:
                Vofimatica= "N/A"
            else:
                Vofimatica= str(dato.equipo.ofimatica)


            if dato.equipo.antivirus is None:
                Vantivirus= "N/A"
            else:
                Vantivirus= str(dato.equipo.antivirus)

            if dato.equipo.estado =="1":
                Vestado= "BUENO"
            if dato.equipo.estado =="2":
                Vestado= "REGULAR"
            if dato.equipo.estado =="3":
                Vestado= "MALO"
           
            if dato.equipo.tipoEquipo is None:
                VtipoEquipo= "N/A"
            else:
                VtipoEquipo= str(dato.equipo.tipoEquipo)

            if dato.equipo.categoria == "1":
                ws.cell(row=cont,column=2).value = str(dato.custodio)
                ws.cell(row=cont,column=3).value = "zona -"+str(dato.custodio.area.distrito.provincia.zona)
                ws.cell(row=cont,column=4).value = str(dato.custodio.area.distrito)
                ws.cell(row=cont,column=5).value = str(dato.custodio.area)
                ws.cell(row=cont,column=6).value = dato.equipo.nombreEquipo
                ws.cell(row=cont,column=7).value = VtipoEquipo
                ws.cell(row=cont,column=8).value = VMarca
                ws.cell(row=cont,column=9).value = Vmodelo
                ws.cell(row=cont,column=10).value =VserMies
                ws.cell(row=cont,column=11).value = VcodMies
                ws.cell(row=cont,column=12).value = Vprocesador
                ws.cell(row=cont,column=13).value = VcapacidadDisco
                ws.cell(row=cont,column=14).value = VcapacidadMemoriaRam
                ws.cell(row=cont,column=15).value = VsistemaOperativo
                ws.cell(row=cont,column=16).value = VdireccionIp
                ws.cell(row=cont,column=17).value = VdireccionMac
                ws.cell(row=cont,column=18).value =Vofimatica
                ws.cell(row=cont,column=19).value =Vantivirus
                ws.cell(row=cont,column=20).value = dato.equipo.softwareAdicional
                ws.cell(row=cont,column=21).value = Vestado
                ws.cell(row=cont,column=22).value = dato.equipo.observacion
                cont = cont + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="inventarioComputadoras.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response





#Nuestra clase hereda de la vista genérica TemplateView
class ImpresorasReporteExcelView(TemplateView):
     
     
      #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las lista de nuestra base de listas

        lista = Custodia.objects.filter(estado=1)
        
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE lista'
        ws['B1'] = 'INVENTARIO  DE IMPRESORAS MIES'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'NOMBRE DEL CUSTODIO'
        ws['C3'] = 'UNIDAD EJECUTORA'
        ws['D3'] = 'DEPENDENCIA'
        ws['E3'] = 'UNIDAD ADMINISTRATIVA'
        ws['F3'] = 'MARCA'
        ws['G3'] = 'MODELO'
        ws['H3'] = 'SERIE'
        ws['I3'] = 'CÓDIGO MIES'
        ws['J3'] = 'TECNOLOGÍA'
        ws['K3'] = 'TIPO'
        ws['L3'] = 'DIRECCIÓN IP'
        ws['M3'] = 'DIRECCIÓN MAC'
        ws['N3'] = 'ESTADO'
        ws['O3'] = 'OBSERVACIONES'
       

   #     ws.merge_cells('R3:Y3')
        #ancho de columna
        ws.column_dimensions['B'].width = 40.0
        ws.column_dimensions['C'].width = 20.0
        ws.column_dimensions['D'].width = 35.0
        ws.column_dimensions['E'].width = 35.0
        ws.column_dimensions['F'].width = 30.0
        ws.column_dimensions['G'].width = 30.0
        ws.column_dimensions['H'].width = 25.0
        ws.column_dimensions['I'].width = 25.0
        ws.column_dimensions['J'].width = 25.0
        ws.column_dimensions['K'].width = 25.0
        ws.column_dimensions['L'].width = 25.0
        ws.column_dimensions['M'].width = 40.0
        ws.column_dimensions['N'].width = 25.0
        ws.column_dimensions['O'].width = 40.0
     
       
        cont=4

        #Recorremos el conjunto de lista y vamos escribiendo cada uno de los listas en las celdas
        
        for dato in lista:
            
            if dato.equipo.codigoMies is None:
                VcodMies= "N/A"
            else:
                VcodMies= dato.equipo.codigoMies
            if dato.equipo.serie is None:
                VserMies= "N/A"
            else:
                VserMies= dato.equipo.serie

            if dato.equipo.direccionIp is None:
                VdireccionIp= "N/A"
            else:
                VdireccionIp= dato.equipo.direccionIp

            if dato.equipo.direccionMac is None:
                VdireccionMac= "N/A"
            else:
                VdireccionMac= dato.equipo.direccionMac
           


            if dato.equipo.marca is None:
                VMarca= "N/A"
            else:
                VMarca= str(dato.equipo.marca)


            if dato.equipo.modelo is None:
                Vmodelo= "N/A"
            else:
                Vmodelo= str(dato.equipo.modelo)



            if dato.equipo.tecnologia is None:
                Vtecnologia= "N/A"
            else:
                Vtecnologia= str(dato.equipo.tecnologia)



            if dato.equipo.tipoImpresora is None:
                VtipoImpresora= "N/A"
            else:
                VtipoImpresora= str(dato.equipo.tipoImpresora)
         

            if dato.equipo.estado =="1":
                Vestado= "BUENO"
            if dato.equipo.estado =="2":
                Vestado= "REGULAR"
            if dato.equipo.estado =="3":
                Vestado= "MALO"
           

            if dato.equipo.categoria == "2":
                ws.cell(row=cont,column=2).value = str(dato.custodio)
                ws.cell(row=cont,column=3).value = "zona -"+str(dato.custodio.area.distrito.provincia.zona)
                ws.cell(row=cont,column=4).value = str(dato.custodio.area.distrito)
                ws.cell(row=cont,column=5).value = str(dato.custodio.area)
                ws.cell(row=cont,column=6).value = VMarca
                ws.cell(row=cont,column=7).value = Vmodelo
                ws.cell(row=cont,column=8).value =VserMies
                ws.cell(row=cont,column=9).value = VcodMies
                ws.cell(row=cont,column=10).value = Vtecnologia
                ws.cell(row=cont,column=11).value = VtipoImpresora
                ws.cell(row=cont,column=12).value = VdireccionIp
                ws.cell(row=cont,column=13).value = VdireccionMac
                ws.cell(row=cont,column=14).value = Vestado
                ws.cell(row=cont,column=15).value = dato.equipo.observacion
                cont = cont + 1
                
        #Establecemos el nombre del archivo
        nombre_archivo ="inventarioImpresoras.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response





#Nuestra clase hereda de la vista genérica TemplateView
class OtrosDispositivosReporteExcelView(TemplateView):
     
     
      #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las lista de nuestra base de listas

        lista = Custodia.objects.filter(estado=1)
        
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE lista'
        ws['B1'] = 'INVENTARIO  OTROS DISPOSITIVOS MIES'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'NOMBRE DEL CUSTODIO'
        ws['C3'] = 'UNIDAD EJECUTORA'
        ws['D3'] = 'DEPENDENCIA'
        ws['E3'] = 'UNIDAD ADMINISTRATIVA'
        ws['F3'] = 'TIPO'
        ws['G3'] = 'MARCA'
        ws['H3'] = 'MODELO'
        ws['I3'] = 'SERIE'
        ws['J3'] = 'CÓDIGO MIES'
        ws['K3'] = 'ESTADO '
        ws['L3'] = 'OBSERVACIONES'
       

   #     ws.merge_cells('R3:Y3')
        #ancho de columna
        ws.column_dimensions['B'].width = 40.0
        ws.column_dimensions['C'].width = 20.0
        ws.column_dimensions['D'].width = 35.0
        ws.column_dimensions['E'].width = 35.0
        ws.column_dimensions['F'].width = 30.0
        ws.column_dimensions['G'].width = 30.0
        ws.column_dimensions['H'].width = 25.0
        ws.column_dimensions['I'].width = 25.0
        ws.column_dimensions['J'].width = 40.0
        ws.column_dimensions['K'].width = 25.0
        ws.column_dimensions['L'].width = 40.0
 
     
       
        cont=4

        #Recorremos el conjunto de lista y vamos escribiendo cada uno de los listas en las celdas
        
        for dato in lista:
            
            if dato.equipo.codigoMies is None:
                VcodMies= "N/A"
            else:
                VcodMies= dato.equipo.codigoMies
            if dato.equipo.serie is None:
                VserMies= "N/A"
            else:
                VserMies= dato.equipo.serie        


            if dato.equipo.marca is None:
                VMarca= "N/A"
            else:
                VMarca= str(dato.equipo.marca)


            if dato.equipo.modelo is None:
                Vmodelo= "N/A"
            else:
                Vmodelo= str(dato.equipo.modelo)


            if dato.equipo.tipoDispositivo is None:
                VtipoDispositivo= "N/A"
            else:
                VtipoDispositivo= str(dato.equipo.tipoDispositivo)         

            if dato.equipo.estado =="1":
                Vestado= "BUENO"
            if dato.equipo.estado =="2":
                Vestado= "REGULAR"
            if dato.equipo.estado =="3":
                Vestado= "MALO"
           

            if dato.equipo.categoria == "3":
                ws.cell(row=cont,column=2).value = str(dato.custodio)
                ws.cell(row=cont,column=3).value = "zona -"+str(dato.custodio.area.distrito.provincia.zona)
                ws.cell(row=cont,column=4).value = str(dato.custodio.area.distrito)
                ws.cell(row=cont,column=5).value = str(dato.custodio.area)
                ws.cell(row=cont,column=6).value = VtipoDispositivo
                ws.cell(row=cont,column=7).value = VMarca
                ws.cell(row=cont,column=8).value = Vmodelo
                ws.cell(row=cont,column=9).value =VserMies
                ws.cell(row=cont,column=10).value = VcodMies
                ws.cell(row=cont,column=11).value = Vestado
                ws.cell(row=cont,column=12).value = dato.equipo.observacion
                cont = cont + 1
                
        #Establecemos el nombre del archivo
        nombre_archivo ="inventarioOtrosDispositivos.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class CustodiaCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'custodia.add_custodia'
    model = Custodia
    form_class = CustodiaForm
    template_name = "custodia/custodia_crear.html"
    success_url = reverse_lazy('custodia_listar')

    def get_context_data(self, **kwargs):   
        context = super().get_context_data(**kwargs)
        #obtengo los equipos que ya han sido asignado a un equipo y estan vigentes
        conCustodia = Custodia.objects.all().exclude( estado=0).values_list('equipo')
        #excluyo los dispositivos que esten con estado malo
        dispositivo = Dispositivo.objects.values_list('id')
        #obtengo los id de los dispositivos que no tienen custodio
        diferente  =dispositivo.difference(conCustodia).order_by('id')
        #filtro para mostrar solo los dispositivos sin custodio
        context["equipos"] = Dispositivo.objects.filter(id__in=diferente)
      
        return context  
    
   
class CustodiaReasignarView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'custodia.add_custodia'
    model = Custodia
    form_class = CustodiaForm
    template_name = "custodia/custodiaReasignar.html"
    success_url = reverse_lazy('custodia_listar')

    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #obtengo los equipos que ya han sido asignado a un equipo y estan vigentes
        conCustodia = Custodia.objects.all().exclude( estado=0).values_list('equipo')
        print (conCustodia)
        #excluyo los dispositivos que esten con estado malo
        dispositivo = Dispositivo.objects.exclude(estado=3).values_list('id')
        print (dispositivo)
        #obtengo los id de los dispositivos que no tienen custodio
        interseccion  =dispositivo.intersection(conCustodia).order_by('id')
        print (interseccion)
        #filtro para mostrar solo los dispositivos sin custodio
        context["equipos"] = Dispositivo.objects.filter(id__in=interseccion)
        #obtener todos los custodios para saber quien es el custodio anterior
        context["custodioAnterior"] = Custodia.objects.exclude( estado=0)
      
        return context  

 

class CustodiaDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'custodia.delete_custodia'
    model = Custodia
    template_name = "custodia/custodia_eliminar.html"
    success_url = reverse_lazy('custodia_listar')
   


class CustodiaUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'custodia.change_custodia'
    model = Custodia
    form_class = CustodiaForm
    template_name = "custodia/custodia_editar.html"
    success_url = reverse_lazy('custodia_listar')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #todos los dispositivos 
        context["equipos"] = Dispositivo.objects.all()
        context["custodioAnterior"] = Custodia.objects.filter(id=self.object.id)
        context["update_equipo"] = Dispositivo.objects.filter(id=self.object.equipo.id)
        context["update_categoria"] = self.object.equipo.categoria
        print(self.object.equipo.categoria)
        return context
    
    
class CustodiaDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'custodia.view_custodia'
    model = Custodia
    template_name = "custodia/custodia_detalle.html"

    