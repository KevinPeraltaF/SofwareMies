from django.shortcuts import render
from django.urls import reverse_lazy
from django.views.generic import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import (CreateView, UpdateView, DeleteView)
from .models import Custodio
from .forms import CustodioForm
from inventario.models import EquipoDetalle
from django.contrib.auth.mixins import PermissionRequiredMixin,LoginRequiredMixin
#Empleado
from inventario.models import EquipoCabecera
#excel
# excel
#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
import csv
from csv import reader

from openpyxl.utils import get_column_letter

# Create your views here.
class CustodioListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'custodio.view_custodio'
    model = Custodio
    template_name = "custodio/custodio_listado.html"

    #def get_queryset(self):
     #   return Custodio.objects.exclude(estado = 0) CustodioReporteExcelView
    
# excel


 #excel 
 
#Nuestra clase hereda de la vista genérica TemplateView
class CustodioReporteExcelView(TemplateView):
     
     
      #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las lista de nuestra base de listas
        lista = Custodio.objects.filter(estado=1)
        
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE lista'
        ws['B1'] = 'INVENTARIO DE TICS'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws['C3'] = 'ZONA'
        ws['D3'] = 'DISTRITO'
        ws['E3'] = 'AREA'
        ws['F3'] = 'CUSTODIO'
        ws['G3'] = 'CATEGORIA'
        ws['H3'] = 'CONDICIÓN'
        ws['I3'] = 'MARCA'
        ws['J3'] = 'MODELO'
        ws['K3'] = 'SERIE'
        ws['L3'] = 'CODIGO MIES'
        ws['M3'] = 'DIRECCIÓN IP'
        ws['N3'] = 'DIRECCIÓN MAC'
        ws['O3'] = 'CAPACIDAD DISCO'
        ws['P3'] = 'CAPACIDAD RAM'
        ws['Q3'] = 'PROCESADOR'
        ws['R3'] = 'PERIFERICOS DEL EQUIPO'
       

        ws.merge_cells('R3:Y3')
        #ancho de columna
        ws.column_dimensions['B'].width = 15.0
        ws.column_dimensions['C'].width = 30.0
        ws.column_dimensions['D'].width = 25.0
        ws.column_dimensions['E'].width = 20.0
        ws.column_dimensions['F'].width = 20.0
        ws.column_dimensions['G'].width = 25.0
        ws.column_dimensions['H'].width = 25.0
        ws.column_dimensions['I'].width = 25.0
        ws.column_dimensions['J'].width = 25.0
        ws.column_dimensions['K'].width = 25.0
        ws.column_dimensions['L'].width = 25.0
        ws.column_dimensions['M'].width = 25.0
        ws.column_dimensions['N'].width = 25.0
        ws.column_dimensions['O'].width = 25.0
        ws.column_dimensions['P'].width = 25.0
        ws.column_dimensions['Q'].width = 25.0
        ws.column_dimensions['R'].width = 25.0
        ws.column_dimensions['S'].width = 25.0
        ws.column_dimensions['T'].width = 25.0
        ws.column_dimensions['U'].width = 25.0
        ws.column_dimensions['V'].width = 30.0
        ws.column_dimensions['W'].width = 30.0
        ws.column_dimensions['X'].width = 30.0
        ws.column_dimensions['Y'].width = 30.0
        
        cont=4
        enumerador = 1
        #Recorremos el conjunto de lista y vamos escribiendo cada uno de los listas en las celdas
        
        for dato in lista:
            contadorDetalle =18
            ws.cell(row=cont,column=2).value = enumerador
            ws.cell(row=cont,column=3).value = dato.custodio.area.distrito.provincia.zona.descripcion
            ws.cell(row=cont,column=4).value = dato.custodio.area.distrito.descripcion
            ws.cell(row=cont,column=5).value = dato.custodio.area.descripcion
            ws.cell(row=cont,column=6).value = dato.custodio.apellidos + ' '+ dato.custodio.nombres
            ws.cell(row=cont,column=7).value = dato.equipo.categoria.descripcion
            ws.cell(row=cont,column=8).value = dato.equipo.condicion
            ws.cell(row=cont,column=9).value = dato.equipo.marca.descripcion
            ws.cell(row=cont,column=10).value = dato.equipo.modelo.descripcion
            ws.cell(row=cont,column=11).value = dato.equipo.serie
            ws.cell(row=cont,column=12).value = dato.equipo.codigoMies
            ws.cell(row=cont,column=13).value = dato.equipo.direccionIp
            ws.cell(row=cont,column=14).value = dato.equipo.direccionMac
            ws.cell(row=cont,column=15).value = dato.equipo.capacidadDisco.descripcion
            ws.cell(row=cont,column=16).value = dato.equipo.capacidadMemoria.descripcion
            ws.cell(row=cont,column=17).value = dato.equipo.capacidadProcesador.descripcion
            listaDetalle = EquipoDetalle.objects.filter(cabeceraDistrito=dato.equipo.id)
            for detalle in listaDetalle:
                ws.cell(row=cont,column=contadorDetalle).value = detalle.periferico.descripcion
                contadorDetalle=contadorDetalle+1
            enumerador = enumerador+1
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteInventarioDistrito.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response



class CustodioCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'custodio.add_custodio'
    model = Custodio
    form_class = CustodioForm
    template_name = "custodio/custodio_crear.html"
    success_url = reverse_lazy('custodio_listar')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #condicion dañado
        context["equipo"] = EquipoCabecera.objects.exclude( condicion=3)
        context["custodioAnterior"] = Custodio.objects.all()
      
        return context
    
   

class CustodioDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'custodio.delete_custodio'
    model = Custodio
    template_name = "custodio/custodio_eliminar.html"
    success_url = reverse_lazy('custodio_listar')
   


class CustodioUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'custodio.change_custodio'
    model = Custodio
    form_class = CustodioForm
    template_name = "custodio/custodio_editar.html"
    success_url = reverse_lazy('custodio_listar')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["equipo"] = EquipoCabecera.objects.all()
        context["custodioAnterior"] = Custodio.objects.all()
      
        return context
    


    
class CustodioDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'custodio.view_custodio'
    model = Custodio
    template_name = "custodio/custodio_detalle.html"

