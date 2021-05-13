from django.shortcuts import render
from django.urls import reverse_lazy
from django.views.generic import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import (CreateView, UpdateView, DeleteView)
from .models import Empleado, Area, Cargo, UnidadAtencion
from .forms import EmpleadoForm, AreaForm, CargoForm, UnidadAtencionForm
from django.contrib.auth.mixins import LoginRequiredMixin,PermissionRequiredMixin
from django.utils.decorators import method_decorator
from functools import wraps
from django.utils.translation import gettext_lazy as _
from django.db.models.deletion import ProtectedError
from django.core.exceptions import PermissionDenied


# excel
#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
import csv
from csv import reader

from openpyxl.utils import get_column_letter


def protected_error_as_api_error():
    """
    Decorator to handle all `ProtectedError` error as API Error,
    which mean, converting from error 500 to error 403.
    """
    def decorator(func):
        @wraps(func)
        def wrapper(request, *args, **kwargs):
            try:
                return func(request, *args, **kwargs)
            except ProtectedError as error:
                raise PermissionDenied(_('Action Denied: The selected object is being used '
                                         'by the system. Deletion not allowed.'))
        return wrapper
    return decorator

# VISTA EMPLEADO.
class EmpleadoListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'empleado.view_empleado'
    model = Empleado
    template_name = "empleado/empleado_listado.html"
    
class EmpleadoCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'empleado.add_empleado'
    model = Empleado
    form_class = EmpleadoForm
    template_name = "empleado/empleado_crear.html"
    success_url = reverse_lazy('empleado_listar')
  


class EmpleadoDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'empleado.delete_empleado'
    model = Empleado
    template_name = "empleado/empleado_eliminar.html"
    success_url = reverse_lazy('empleado_listar')



class EmpleadoUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'empleado.change_empleado'
    model = Empleado
    form_class = EmpleadoForm
    template_name = "empleado/empleado_editar.html"
    success_url = reverse_lazy('empleado_listar')

    
class EmpleadoDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'empleado.view_empleado'
    model = Empleado
    template_name = "empleado/empleado_detalle.html"

    #excel empleado




#Nuestra clase hereda de la vista genérica TemplateView
class EmpleadoExcelListView(TemplateView):
     
    #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las lista de nuestra base de listas
        lista = Empleado.objects.all()
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE lista'
        ws['B1'] = 'EMPLEADOS DEL DISTRITO'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'AREA'
        ws['C3'] = 'CARGO'
        ws['D3'] = 'NOMBRE'
        ws['E3'] = 'CÉDULA'
        ws['F3'] = 'CORREO'
        ws['G3'] = 'GENERO'
        ws['H3'] = 'TELEFONO'
        ws['I3'] = 'UNIDAD DE ATENCIÓN'
        ws['J3'] = 'ESTADO DEL EMPLEADO'
      


        #ancho de columna
        ws.column_dimensions['B'].width = 30.0
        ws.column_dimensions['C'].width = 60.0
        ws.column_dimensions['D'].width = 40.0
        ws.column_dimensions['E'].width = 15.0
        ws.column_dimensions['F'].width = 40.0
        ws.column_dimensions['G'].width = 15.0
        ws.column_dimensions['H'].width = 15.0
        ws.column_dimensions['I'].width = 40.0
        ws.column_dimensions['J'].width = 25.0


        cont=4
        
        #Recorremos el conjunto de lista y vamos escribiendo cada uno de los listas en las celdas
        for dato in lista:

            if dato.unidadAtencion is None:
                DatounidadAtencion = 'N/A'
            else:
                DatounidadAtencion = dato.unidadAtencion.descripcion +"-"+dato.unidadAtencion.codigo
               

            if dato.telefono is None:
                DatoTelefono = 'N/A'
            else:
                DatoTelefono = dato.telefono

            if dato.genero == "1":
                DatoGenero = 'MASCULINO'
            if dato.genero == "2":
                DatoGenero = 'FEMENINO'
            if dato.genero == "3":
                DatoGenero = 'GLBTI'

            if dato.estado:
                datoEstado = 'Activo'

            if not dato.estado:
                datoEstado = 'No activo'
          
             
            ws.cell(row=cont,column=2).value = dato.area.descripcion
            ws.cell(row=cont,column=3).value =  dato.cargo.descripcion
            ws.cell(row=cont,column=4).value =(dato.apellidos +" " +dato.nombres)
            ws.cell(row=cont,column=5).value = dato.cedula
            ws.cell(row=cont,column=6).value = dato.correo
            ws.cell(row=cont,column=7).value = DatoGenero
            ws.cell(row=cont,column=8).value = DatoTelefono
            ws.cell(row=cont,column=9).value = DatounidadAtencion
            ws.cell(row=cont,column=10).value = datoEstado
 
         
            cont = cont + 1
    

        #Establecemos el nombre del archivo
        nombre_archivo ="ListadoEmpleadoDistrito.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response



# VISTA AREA.
class AreaListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'empleado.view_area'
    model = Area
    template_name = "empleado/area_listado.html"

    

class AreaCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'empleado.add_area'
    model = Area
    form_class = AreaForm
    template_name = "empleado/area_crear.html"
    success_url = reverse_lazy('area_listar')

  

class AreaDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'empleado.delete_area'
    model = Area
    template_name = "empleado/area_eliminar.html"
    success_url = reverse_lazy('area_listar')
    
    @method_decorator(protected_error_as_api_error())
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    

   


class AreaUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'empleado.change_area'
    model = Area
    form_class = AreaForm
    template_name = "empleado/area_editar.html"
    success_url = reverse_lazy('area_listar')


    
class AreaDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'empleado.view_area'
    model = Area
    template_name = "empleado/area_detalle.html"

# VISTA CARGO.
class CargoListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'empleado.view_cargo'
    model = Cargo
    template_name = "empleado/cargo_listado.html"

    
class CargoCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'empleado.add_cargo'
    model = Cargo
    form_class = CargoForm
    template_name = "empleado/cargo_crear.html"
    success_url = reverse_lazy('cargo_listar')


class CargoDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'empleado.delete_cargo'
    model = Cargo
    template_name = "empleado/cargo_eliminar.html"
    success_url = reverse_lazy('cargo_listar')


class CargoUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'empleado.change_cargo'
    model = Cargo
    form_class = CargoForm
    template_name = "empleado/cargo_editar.html"
    success_url = reverse_lazy('cargo_listar')

    
class CargoDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'empleado.view_cargo'
    model = Cargo
    template_name = "empleado/cargo_detalle.html"


# VISTA UNIDAD DE ATENCION.
class UnidadListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'empleado.view_unidadatencion'
    model = UnidadAtencion
    template_name = "empleado/unidad_listado.html"

    
class UnidadCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'empleado.add_unidadatencion'
    model = UnidadAtencion
    form_class = UnidadAtencionForm
    template_name = "empleado/unidad_crear.html"
    success_url = reverse_lazy('unidad_listar')

class UnidadDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'empleado.delete_unidadatencion'
    model = UnidadAtencion
    template_name = "empleado/unidad_eliminar.html"
    success_url = reverse_lazy('unidad_listar')


class UnidadUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'empleado.change_unidadatencion'
    model = UnidadAtencion
    form_class = UnidadAtencionForm
    template_name = "empleado/unidad_editar.html"
    success_url = reverse_lazy('unidad_listar')

    
class UnidadDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'empleado.view_unidadatencion'
    model = UnidadAtencion
    template_name = "empleado/unidad_detalle.html"
