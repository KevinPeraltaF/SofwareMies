from django.shortcuts import render
from django.urls import reverse_lazy
from django.views.generic import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import (CreateView, UpdateView, DeleteView)
from django.contrib.auth.mixins import LoginRequiredMixin,PermissionRequiredMixin
from .models import InventarioTics, Marca, Modelo, Condicion, Categoria, CapacidadDisco, CapacidadMemoriaRam, Procesador,\
    EquipoCabecera, EquipoDetalle
from .forms import InvTicsForm, MarcaForm, ModeloForm, CategoriaForm, CondicionForm, CapacidadDiscoForm, CapacidadMemoriaRamForm, ProcesadorForm,\
    EquipoCabeceraForm, EquipoDetalleForm,EquipoCabDetalleForm
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponseRedirect, JsonResponse
import json
from django.db import transaction
from django.forms import model_to_dict
# excel
#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
import csv
from csv import reader

from openpyxl.utils import get_column_letter

# Create your views here.
#------------------MARCA--------------------------------------
class MarcaListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'inventario.view_marca'
    model = Marca
    template_name = "inventario/marca_listado.html"


class MarcaCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'inventario.add_marca'
    model = Marca
    form_class = MarcaForm
    template_name= "inventario/marca_crear.html"
    success_url = reverse_lazy('marca_listar')

class MarcaDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'inventario.delete_marca'
    model = Marca
    template_name = "inventario/marca_eliminar.html"
    success_url = reverse_lazy('marca_listar')
 

class MarcaUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'inventario.change_marca'
    model = Marca
    form_class = MarcaForm
    template_name = "inventario/marca_editar.html"
    success_url = reverse_lazy('marca_listar')
   
class MarcaDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'inventario.view_marca'
    model = Marca
    template_name = "inventario/marca_detalle.html"
#------------------MODELO--------------------------------------
class ModeloListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'inventario.view_modelo'
    model = Modelo
    template_name = "inventario/modelo_listado.html"


class ModeloCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'inventario.add_modelo'
    model = Modelo
    form_class = ModeloForm
    template_name= "inventario/modelo_crear.html"
    success_url = reverse_lazy('modelo_listar')
 
class ModeloDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'inventario.delete_modelo'
    model = Modelo
    template_name = "inventario/modelo_eliminar.html"
    success_url = reverse_lazy('modelo_listar')
 

class ModeloUpdateView(LoginRequiredMixin,PermissionRequiredMixin, UpdateView):
    permission_required = 'inventario.change_modelo'
    model = Modelo
    form_class = ModeloForm
    template_name = "inventario/modelo_editar.html"
    success_url = reverse_lazy('modelo_listar')
 
class ModeloDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'inventario.view_modelo'
    model = Modelo
    template_name = "inventario/modelo_detalle.html"
#------------------CATEGORIA--------------------------------------
class CategoriaListView(LoginRequiredMixin,PermissionRequiredMixin, ListView):
    permission_required = 'inventario.view_categoria'
    model = Categoria
    template_name = "inventario/categoria_listado.html"

  
class CategoriaCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'inventario.add_categoria'
    model = Categoria
    form_class = CategoriaForm
    template_name= "inventario/categoria_crear.html"
    success_url = reverse_lazy('categoria_listar')
    
class CategoriaDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'inventario.delete_categoria'
    model = Categoria
    template_name = "inventario/categoria_eliminar.html"
    success_url = reverse_lazy('categoria_listar')
 


class CategoriaUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'inventario.change_categoria'
    model = Categoria
    form_class = CategoriaForm
    template_name = "inventario/categoria_editar.html"
    success_url = reverse_lazy('categoria_listar')

class CategoriaDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'inventario.view_categoria'
    model = Categoria
    template_name = "inventario/categoria_detalle.html"
#------------------CONDICION--------------------------------------
class CondicionListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'inventario.view_condicion'
    model = Condicion
    template_name = "inventario/condicion_listado.html"

 
class CondicionCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'inventario.add_condicion'
    model = Condicion
    form_class = CondicionForm
    template_name= "inventario/condicion_crear.html"
    success_url = reverse_lazy('condicion_listar')

class CondicionDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'inventario.delete_condicion'
    model = Condicion
    template_name = "inventario/condicion_eliminar.html"
    success_url = reverse_lazy('condicion_listar')
  

class CondicionUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'inventario.change_condicion'
    model = Condicion
    form_class = CondicionForm
    template_name = "inventario/condicion_editar.html"
    success_url = reverse_lazy('condicion_listar')
  
class CondicionDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'inventario.view_condicion'
    model = Condicion
    template_name = "inventario/condicion_detalle.html"
#------------------INVENTARIO TICS--------------------------------------
class InvTicsListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'inventario.view_inventariotics'
    model = InventarioTics
    template_name = "inventario/inv_tics_listado.html"

 #excel 
 
#Nuestra clase hereda de la vista genérica TemplateView
class InvTicsExcelListView(TemplateView):
     
    #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las lista de nuestra base de listas
        lista = InventarioTics.objects.all()
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE lista'
        ws['B1'] = 'INVENTARIO DE TICS'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'FECHA INGRESO'
        ws['C3'] = 'RESPOSABLE'
        ws['D3'] = 'ITEMS'
        ws['E3'] = 'CODIGO MIES'
        ws['F3'] = 'SERIE'
        ws['G3'] = 'CATEGORIA'
        ws['H3'] = 'CANTIDAD'
        ws['I3'] = 'MARCA'
        ws['J3'] = 'MODELO'
        ws['K3'] = 'UBICACIÓN'
        ws['L3'] = 'CONDICIÓN'


        #ancho de columna
        ws.column_dimensions['B'].width = 15.0
        ws.column_dimensions['C'].width = 30.0
        ws.column_dimensions['D'].width = 25.0
        ws.column_dimensions['E'].width = 20.0
        ws.column_dimensions['F'].width = 20.0
        ws.column_dimensions['G'].width = 25.0
        ws.column_dimensions['H'].width = 15.0
        ws.column_dimensions['I'].width = 25.0
        ws.column_dimensions['J'].width = 25.0
        ws.column_dimensions['K'].width = 25.0
        ws.column_dimensions['L'].width = 15.0

        cont=4
        
        #Recorremos el conjunto de lista y vamos escribiendo cada uno de los listas en las celdas
        for dato in lista:

            if dato.codigoMies is None:
                DatoCodigoMies = 'N/A'
            else:
                DatoCodigoMies = dato.codigoMies
               

            if dato.serie is None:
                DatoSerie = 'N/A'
            else:
                DatoSerie = dato.serie
               
            if dato.marca is None:
                DatoMarca = 'N/A'
            else:
                DatoMarca = dato.marca.descripcion
               
            if dato.modelo is None:
                DatoModelo = 'N/A'
            else:
                 DatoModelo = dato.marca.descripcion
            
            ws.cell(row=cont,column=2).value = dato.fechaIngreso
            ws.cell(row=cont,column=3).value = dato.responsable.apellidos
            ws.cell(row=cont,column=4).value = dato.descripcion
            ws.cell(row=cont,column=5).value = DatoCodigoMies
            ws.cell(row=cont,column=6).value = DatoSerie
            ws.cell(row=cont,column=7).value = dato.categoria.descripcion
            ws.cell(row=cont,column=8).value = dato.cantidad
            ws.cell(row=cont,column=9).value = DatoMarca
            ws.cell(row=cont,column=10).value = DatoModelo
            ws.cell(row=cont,column=11).value = dato.ubicacion.descripcion
            ws.cell(row=cont,column=12).value = dato.condicion.descripcion
         
            cont = cont + 1
    

        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteInventarioTics.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class InvTicsCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'inventario.add_inventariotics'
    model = InventarioTics
    form_class = InvTicsForm
    template_name= "inventario/inv_tics_crear.html"
    success_url = reverse_lazy('inv_tics_listar')
    

class InvTicsDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'inventario.delete_inventariotics'
    model = InventarioTics
    template_name = "inventario/inv_tics_eliminar.html"
    success_url = reverse_lazy('inv_tics_listar')
    
    
class InvTicsUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'inventario.change_inventariotics'
    model = InventarioTics
    form_class = InvTicsForm
    template_name = "inventario/inv_tics_editar.html"
    success_url = reverse_lazy('inv_tics_listar')
    
        
class InvTicsDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'inventario.view_inventariotics'
    model = InventarioTics
    template_name = "inventario/inv_tics_detalle.html"
#------------------CAPACIDAD DISCO--------------------------------------
class CapacidadDiscoListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'inventario.view_capacidaddisco'
    model = CapacidadDisco
    template_name = "inventario/capacidad_disco_listado.html"

   
    
class CapacidadDiscoCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'inventario.add_capacidaddisco'
    model = CapacidadDisco
    form_class = CapacidadDiscoForm
    template_name= "inventario/capacidad_disco_crear.html"
    success_url = reverse_lazy('capacidad_disco_listar')
  

class CapacidadDiscoDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'inventario.delete_capacidaddisco'
    model = CapacidadDisco
    template_name = "inventario/capacidad_discos_eliminar.html"
    success_url = reverse_lazy('capacidad_disco_listar')
    
    
class CapacidadDiscoUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'inventario.change_capacidaddisco'
    model = CapacidadDisco
    form_class = CapacidadDiscoForm
    template_name = "inventario/capacidad_disco_editar.html"
    success_url = reverse_lazy('capacidad_disco_listar')
   
        
class CapacidadDiscoDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'inventario.view_capacidaddisco'
    model = CapacidadDisco
    template_name = "inventario/capacidad_disco_detalle.html"
#------------------CAPACIDAD MEMORIA RAM--------------------------------------
class CapacidadMemoriaRamListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'inventario.view_capacidadmemoriaram'
    model = CapacidadMemoriaRam
    template_name = "inventario/capacidad_memoria_ram_listado.html"

    
    
class CapacidadMemoriaRamCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'inventario.add_capacidadmemoriaram'
    model = CapacidadMemoriaRam
    form_class = CapacidadMemoriaRamForm
    template_name= "inventario/capacidad_memoria_ram_crear.html"
    success_url = reverse_lazy('capacidad_memoria_ram_listar')
    

class CapacidadMemoriaRamDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'inventario.delete_capacidadmemoriaram'
    model = CapacidadMemoriaRam
    template_name = "inventario/capacidad_memoria_ram_eliminar.html"
    success_url = reverse_lazy('capacidad_memoria_ram_listar')
  
    
class CapacidadMemoriaRamUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'inventario.change_capacidadmemoriaram'
    model = CapacidadMemoriaRam
    form_class = CapacidadMemoriaRamForm
    template_name = "inventario/capacidad_memoria_ram_editar.html"
    success_url = reverse_lazy('capacidad_memoria_ram_listar')
 
        
class CapacidadMemoriaRamDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'inventario.view_capacidadmemoriaram'
    model = CapacidadMemoriaRam
    template_name = "inventario/capacidad_memoria_ram_detalle.html"
#------------------PROCESADOR--------------------------------------
class ProcesadorListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'inventario.view_procesador'
    model = Procesador
    template_name = "inventario/procesador_listado.html"

    
class ProcesadorCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'inventario.add_procesador'
    model = Procesador
    form_class = ProcesadorForm
    template_name= "inventario/procesador_crear.html"
    success_url = reverse_lazy('procesador_listar')


class ProcesadorDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'inventario.delete_procesador'
    model = Procesador
    template_name = "inventario/procesador_eliminar.html"
    success_url = reverse_lazy('procesador_listar')
    
    
class ProcesadorUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'inventario.change_procesador'
    model = Procesador
    form_class = ProcesadorForm
    template_name = "inventario/procesador_editar.html"
    success_url = reverse_lazy('procesador_listar')
    
        
class ProcesadorDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'inventario.view_procesador'
    model = Procesador
    template_name = "inventario/procesador_detalle.html"
#------------------CABECERA-DETALLE-EQUIPO--------------------------------------

class EquipoListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'inventario.view_equipocabecera'
    model = EquipoCabecera
    template_name = "inventario/equipo_listado.html"
    @method_decorator(csrf_exempt)
    def dispatch(self,request,*args, **kwargs):
        return super().dispatch(request,*args, **kwargs)

        
class EquipoCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'inventario.add_equipocabecera'
    model = EquipoCabecera
    form_class = EquipoCabeceraForm
    template_name= "inventario/equipo_crear.html"
    success_url = reverse_lazy('equipo_listar')
    def get(self, request,*args,**kwargs):
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_form = EquipoCabDetalleForm()
        return self.render_to_response(
            self.get_context_data(form=form, detalle_form=detalle_form)
        )
    
    def post(self, request,*args, **kwargs):
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_form = EquipoCabDetalleForm(self.request.POST)
        if (form.is_valid() and detalle_form.is_valid()):
            return self.form_valid(form, detalle_form)
        else:
            return self.form_invalid(form, detalle_form)

    def form_valid(self, form, detalle_form):
        cabecera = form.save()
        detalle_form.instance = cabecera
        detalle_form.save()
        return HttpResponseRedirect(self.success_url)

    def form_invalid(self, form, detalle_form):
        return self.render_to_response(
            self.get_context_data(form = form, detalle_form=detalle_form)
        )
class EquipoUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'inventario.change_equipocabecera'
    model = EquipoCabecera
    form_class = EquipoCabeceraForm
    template_name= "inventario/equipo_editar.html"
    success_url = reverse_lazy('equipo_listar')
    def get(self, request,*args,**kwargs):
        self.object = self.get_object()
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_form = EquipoCabDetalleForm(instance = self.object)
        return self.render_to_response(
            self.get_context_data(form=form, detalle_form=detalle_form)
        )
    
    def post(self, request,*args, **kwargs):
        self.object = self.get_object()
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_form = EquipoCabDetalleForm(self.request.POST,instance=self.get_object())
        if (form.is_valid() and detalle_form.is_valid()):
            return self.form_valid(form, detalle_form)
        else:
            return self.form_invalid(form, detalle_form)

    def form_valid(self, form, detalle_form):
        cabecera = form.save()
        detalle_form.instance = cabecera
        detalle_form.save()
        return HttpResponseRedirect(self.success_url)

    def form_invalid(self, form, detalle_form):
        return self.render_to_response(
            self.get_context_data(form = form, detalle_form=detalle_form)
        )   
class EquipoDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'inventario.view_equipocabecera'
    model = EquipoCabecera
    template_name = "inventario/equipo_detalle.html"
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
     
        context['items'] = EquipoDetalle.objects.filter(cabeceraDistrito=self.object.id)
        return context
class EquipoDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'inventario.delete_equipocabecera'
    model = EquipoCabecera
    template_name = "inventario/equipo_eliminar.html"
    success_url = reverse_lazy('equipo_listar')




class EquipoReporteExcelView(TemplateView):
     
     
      #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las lista de nuestra base de listas
        lista = EquipoCabecera.objects.all()
        
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE lista'
        ws['B1'] = 'INVENTARIO DE EQUIPOS'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:N1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws['C3'] = 'FECHA INGRESO'
        ws['D3'] = 'CATEGORIA'
        ws['E3'] = 'MARCA'
        ws['F3'] = 'MODELO'
        ws['G3'] = 'CONDICIÓN'
        ws['H3'] = 'SERIE'
        ws['I3'] = 'CODIGO MIES'
        ws['J3'] = 'DIRECCIÓN IP'
        ws['K3'] = 'DIRECCIÓN MAC'
        ws['L3'] = 'CAPACIDAD DISCO'
        ws['M3'] = 'CAPACIDAD RAM'
        ws['N3'] = 'PROCESADOR'
        ws['O3'] = 'PERIFERICOS DEL EQUIPO'
       

        ws.merge_cells('O3:Y3')
        #ancho de columna
        ws.column_dimensions['A'].width = 15.0
        ws.column_dimensions['B'].width = 25.0
        ws.column_dimensions['C'].width = 25.0
        ws.column_dimensions['D'].width = 25.0
        ws.column_dimensions['E'].width = 25.0
        ws.column_dimensions['F'].width = 25.0
        ws.column_dimensions['G'].width = 25.0
        ws.column_dimensions['H'].width = 25.0
        ws.column_dimensions['I'].width = 25.0
        ws.column_dimensions['J'].width = 25.0
        ws.column_dimensions['K'].width = 25.0
        ws.column_dimensions['L'].width = 25.0
        ws.column_dimensions['M'].width = 25.0
        ws.column_dimensions['N'].width = 25.0
        ws.column_dimensions['O'].width = 25.0
        
        cont=4
        enumerador = 1
        #Recorremos el conjunto de lista y vamos escribiendo cada uno de los listas en las celdas
        
        for dato in lista:

            contadorDetalle =15

            if dato.codigoMies is None:
                DatoCodigoMies = 'N/A'
            else:
                DatoCodigoMies = dato.codigoMies
               

            if dato.serie is None:
                DatoSerie = 'N/A'
            else:
                DatoSerie = dato.serie
               
            if dato.marca is None:
                DatoMarca = 'N/A'
            else:
                DatoMarca = dato.marca.descripcion
               
            if dato.modelo is None:
                DatoModelo = 'N/A'
            else:
                DatoModelo = dato.marca.descripcion

            if dato.direccionIp is None:
                DatoiP = 'N/A'
            else:
                DatoiP =dato.direccionIp
                
            if dato.direccionMac is None:
                DatoMac = 'N/A'
            else:
                DatoMac =dato.direccionMac

            ws.cell(row=cont,column=2).value = enumerador
            ws.cell(row=cont,column=3).value = dato.fechaIngreso
            ws.cell(row=cont,column=4).value = dato.categoria.descripcion
            ws.cell(row=cont,column=5).value = DatoMarca
            ws.cell(row=cont,column=6).value = DatoModelo
            ws.cell(row=cont,column=7).value = dato.condicion.descripcion
            ws.cell(row=cont,column=8).value = DatoSerie
            ws.cell(row=cont,column=9).value = DatoCodigoMies 
            ws.cell(row=cont,column=10).value = DatoiP
            ws.cell(row=cont,column=11).value = DatoMac
            ws.cell(row=cont,column=12).value = dato.capacidadDisco.descripcion
            ws.cell(row=cont,column=13).value = dato.capacidadMemoria.descripcion
            ws.cell(row=cont,column=14).value = dato.capacidadProcesador.descripcion
 
            listaDetalle = EquipoDetalle.objects.filter(cabeceraDistrito=dato.id)
            for detalle in listaDetalle:
                ws.cell(row=cont,column=contadorDetalle).value = detalle.periferico.descripcion
                contadorDetalle=contadorDetalle+1
            enumerador = enumerador+1
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteEquipo.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


    