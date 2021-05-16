from django.shortcuts import render
from django.urls import reverse_lazy
from django.views.generic import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import (CreateView, UpdateView, DeleteView)
from .models import Pasante, Carrera, Universidad
from .forms import PasanteForm, CarreraForm, UniversidadForm
from django.contrib.auth.mixins import LoginRequiredMixin,PermissionRequiredMixin
from django.utils.decorators import method_decorator
# excel
#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
import csv
from csv import reader

from openpyxl.utils import get_column_letter

# VISTA PASANTE.


#Nuestra clase hereda de la vista genérica TemplateView
class PasanteExcelListView(TemplateView):
     
    #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las lista de nuestra base de listas
        lista = Pasante.objects.all()
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE lista'
        ws['B1'] = 'PASANTES VINCULADOS MIES'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:L1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws['C3'] = 'NOMBRES'
        ws['D3'] = 'UNIVERSIDAD'
        ws['E3'] = 'CARRERA'
        ws['F3'] = 'CÉDULA'
        ws['G3'] = 'CELULAR'
        ws['H3'] = 'TUTOR PROFESIONAL'
        ws['I3'] = 'HORAS DIARIAS'
        ws['J3'] = 'FECHA INICIO'
        ws['K3'] = 'FECHA FIN'
        ws['L3'] = 'ESTADO'
      


        #ancho de columna
        ws.column_dimensions['B'].width = 10.0
        ws.column_dimensions['C'].width = 60.0
        ws.column_dimensions['D'].width = 40.0
        ws.column_dimensions['E'].width = 60.0
        ws.column_dimensions['F'].width = 30.0
        ws.column_dimensions['G'].width = 30.0
        ws.column_dimensions['H'].width = 50.0
        ws.column_dimensions['I'].width = 30.0
        ws.column_dimensions['J'].width = 20.0
        ws.column_dimensions['K'].width = 20.0
        ws.column_dimensions['L'].width = 30.0
        cont=4
        enumerador =1
        #Recorremos el conjunto de lista y vamos escribiendo cada uno de los listas en las celdas
        for dato in lista:
            if dato.estado:
                datoEstado = 'FINALIZADO'

            if not dato.estado:
                datoEstado = 'EN FUNCIONES'
          
            if dato.fecha_fin is None:
                datoFin = 'N/A'
            else:
                datoFin = dato.fecha_fin
             
            ws.cell(row=cont,column=2).value = enumerador
            ws.cell(row=cont,column=3).value = (dato.apellidos +" "+ dato.nombres)
            ws.cell(row=cont,column=4).value = dato.carrera.universidad.descripcion
            ws.cell(row=cont,column=5).value = dato.carrera.descripcion
            ws.cell(row=cont,column=6).value = dato.cedula
            ws.cell(row=cont,column=7).value = dato.telefono
            ws.cell(row=cont,column=8).value = (dato.tutor_profesional.apellidos+" "+dato.tutor_profesional.nombres )
            ws.cell(row=cont,column=9).value = dato.horas_diarias
            ws.cell(row=cont,column=10).value = dato.fecha_inicio
            ws.cell(row=cont,column=11).value = datoFin
            ws.cell(row=cont,column=12).value = datoEstado
         
            cont = cont + 1
            enumerador=enumerador + 1
    

        #Establecemos el nombre del archivo
        nombre_archivo ="ListadoPasantes.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class PasanteListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'practica.view_pasante'
    model = Pasante
    template_name = "practica/pasante_listado.html"


    
class PasanteCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'practica.add_pasante'
    model = Pasante
    form_class = PasanteForm
    template_name = "practica/pasante_crear.html"
    success_url = reverse_lazy('pasante_listar')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["universidades"] = Universidad.objects.all()
        context["carreras"] = Carrera.objects.all()
        return context
    


class PasanteDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'practica.delete_pasante'
    model = Pasante
    template_name = "practica/pasante_eliminar.html"
    success_url = reverse_lazy('pasante_listar')
   


class PasanteUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'practica.change_pasante'
    model = Pasante
    form_class = PasanteForm
    template_name = "practica/pasante_editar.html"
    success_url = reverse_lazy('pasante_listar')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["universidades"] = Universidad.objects.all()
        context["carreras"] = Carrera.objects.all()
        context["update_carrera"] = self.object.carrera.id
        context["update_universidad"] = self.object.carrera.universidad.id
        return context
 

    
class PasanteDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'practica.view_pasante'
    model = Pasante
    template_name = "practica/pasante_detalle.html"

# VISTA CARRERA.
class CarreraListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'practica.view_carrera'
    model = Carrera
    template_name = "practica/carrera_listado.html"


    
class CarreraCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'practica.add_carrera'
    model = Carrera
    form_class = CarreraForm
    template_name = "practica/carrera_crear.html"
    success_url = reverse_lazy('carrera_listar')
    


class CarreraDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'practica.delete_carrera'
    model = Carrera
    template_name = "practica/carrera_eliminar.html"
    success_url = reverse_lazy('carrera_listar')
   


class CarreraUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'practica.change_carrera'
    model = Carrera
    form_class = CarreraForm
    template_name = "practica/carrera_editar.html"
    success_url = reverse_lazy('carrera_listar')
 

    
class CarreraDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'practica.view_carrera'
    model = Carrera
    template_name = "practica/carrera_detalle.html"

# VISTA UNIVERSIDAD.
class UniversidadListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'practica.view_universidad'
    model = Universidad
    template_name = "practica/universidad_listado.html"

 
    
class UniversidadCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'practica.add_universidad'
    model = Universidad
    form_class = UniversidadForm
    template_name = "practica/universidad_crear.html"
    success_url = reverse_lazy('universidad_listar')
 

class UniversidadDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'practica.delete_universidad'
    model = Universidad
    template_name = "practica/universidad_eliminar.html"
    success_url = reverse_lazy('universidad_listar')
  

class UniversidadUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'practica.change_universidad'
    model = Universidad
    form_class = UniversidadForm
    template_name = "practica/universidad_editar.html"
    success_url = reverse_lazy('universidad_listar')
  
    
class UniversidadDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'practica.view_universidad'
    model = Universidad
    template_name = "practica/universidad_detalle.html"


