from django.shortcuts import render
from django.urls import reverse_lazy
from django.views.generic import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import (CreateView, UpdateView, DeleteView)
from .models import AccesoRed
from .forms import AccesoRedForm

from django.contrib.auth.mixins import PermissionRequiredMixin,LoginRequiredMixin
from django.http import HttpRequest
from django.utils.decorators import method_decorator
# excel
#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
import csv
from csv import reader

from openpyxl.utils import get_column_letter

# Create your views here.
class AccesoRedListView(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = 'red.view_accesored'
    model = AccesoRed
    template_name = "red/acceso_red_listado.html"
    

class AccesoRedCreateView(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = 'red.add_accesored'
    model = AccesoRed
    form_class = AccesoRedForm
    template_name = "red/acceso_red_crear.html"
    success_url = reverse_lazy('accesoRed_listar')
   

class AccesoRedDeleteView(LoginRequiredMixin,PermissionRequiredMixin,DeleteView):
    permission_required = 'red.delete_accesored'
    model = AccesoRed
    template_name = "red/acceso_red_eliminar.html"
    success_url = reverse_lazy('accesoRed_listar')
   


class AccesoRedUpdateView(LoginRequiredMixin,PermissionRequiredMixin,UpdateView):
    permission_required = 'red.change_accesored'
    model = AccesoRed
    form_class = AccesoRedForm
    template_name = "red/acceso_red_editar.html"
    success_url = reverse_lazy('accesoRed_listar')


    
class AccesoRedDetailView(LoginRequiredMixin,PermissionRequiredMixin,DetailView):
    permission_required = 'red.view_accesored'
    model = AccesoRed
    template_name = "red/acceso_red_detalle.html"

#Nuestra clase hereda de la vista genérica TemplateView
class AccesoRedExcelListView(TemplateView):
     
    #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las lista de nuestra base de listas
        lista = AccesoRed.objects.all()
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE lista'
        ws['B1'] = 'DISPOSITIVOS CONECTADOS A LA RED WIFI'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:G1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws['C3'] = 'USUARIO'
        ws['D3'] = 'DIRECCIÓN IP'
        ws['E3'] = 'DIRECCIÓN MAC'
        ws['F3'] = 'ESTADO'
        ws['G3'] = 'OBSERVACIÓN'
    
      


        #ancho de columna
        ws.column_dimensions['B'].width = 10.0
        ws.column_dimensions['C'].width = 60.0
        ws.column_dimensions['D'].width = 40.0
        ws.column_dimensions['E'].width = 40.0
        ws.column_dimensions['F'].width = 30.0
        ws.column_dimensions['G'].width = 30.0

        cont=4
        enumerador =1
        #Recorremos el conjunto de lista y vamos escribiendo cada uno de los listas en las celdas
        for dato in lista:
            if dato.estado:
                datoEstado = 'CONECTADO'

            if not dato.estado:
                datoEstado = 'NO CONECTADO'
          
        
             
            ws.cell(row=cont,column=2).value = enumerador
            ws.cell(row=cont,column=3).value = dato.usuario
            ws.cell(row=cont,column=4).value = dato.direccion_ip
            ws.cell(row=cont,column=5).value = dato.direccion_mac
            ws.cell(row=cont,column=6).value = datoEstado
            ws.cell(row=cont,column=7).value = dato.observacion
         
            cont = cont + 1
            enumerador=enumerador + 1
    

        #Establecemos el nombre del archivo
        nombre_archivo ="ListadoConectadosRed.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response




